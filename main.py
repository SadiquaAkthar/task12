import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime

class OrangeHRM:
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 10)

    def login(self, username, password):
        self.driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        self.driver.find_element(By.NAME, "username").send_keys(username)
        self.driver.find_element(By.NAME, "password").send_keys(password)
        self.driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button').click()
        self.wait.until(EC.url_contains("dashboard"))

    def is_login_successful(self):
        return "dashboard" in self.driver.current_url

@pytest.mark.parametrize("username, password", [
    ("Admin", "admin123"),
    ("user2", "pass2"),
    ("user3", "pass3"),
    ("user4", "pass4"),
    ("user5", "pass5")
])
def test_login_and_record_result(username, password):
    driver = webdriver.Chrome()  # Update this with the appropriate webdriver
    Orange_HRM = OrangeHRM(driver)

    try:
        Orange_HRM.login(username, password)
        test_result = "Passed" if Orange_HRM.is_login_successful() else "Failed"
    except Exception as e:
        print(f"Error during login: {e}")
        test_result = "Failed"
    finally:
        driver.quit()

    # Record test results in Excel file
    record_test_result(username, password, test_result)

def record_test_result(username, password, test_result):
    excel_file = "test_results.xlsx"
    wb = openpyxl.load_workbook(excel_file)

    sheet = wb.active
    row = sheet.max_row + 1

    test_id = row - 1
    date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tester_name = "Your Tester Name"  # Update this with the actual tester name

    sheet.cell(row=row, column=1, value=test_id)
    sheet.cell(row=row, column=2, value=username)
    sheet.cell(row=row, column=3, value=password)
    sheet.cell(row=row, column=4, value=date_time)
    sheet.cell(row=row, column=5, value=tester_name)
    sheet.cell(row=row, column=6, value=test_result)

    wb.save(excel_file)

if __name__ == "__main__":
    pytest.main()
